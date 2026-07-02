#!/usr/bin/env python3
"""Write extracted listings into the 'Account Procurement' workbook.

Fills the exact 23 client columns for the fields we can extract deterministically.
Leaves the two geo columns BLANK on purpose (a human decides). Appends
helper columns that carry the geo clues and review flags so the analyst has
everything in one row.
"""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# The client's 23 columns, in their exact order. Do not reorder.
TEMPLATE_COLUMNS = [
    "Date of Finding", "URL", "Vendor Type", "Vendor Name", "Ecosystem (Geo)",
    "Account Type", "Details", "Age", "Verified", "Farmed", "Spend (if applicable)",
    "Payment Attached", "PIN", "History", "Geo of Acc.", "Transfer Method", "Cost",
    "No. Available", "Payment Method (if detailed)",
    "Screenshot 1", "Screenshot 2", "Screenshot 3", "Screenshot 4",
]

# Tool-added helper columns (clues + provenance). Prefixed so they read as ours.
HELPER_COLUMNS = [
    "» Confidence", "» Needs Review", "» Clue: Language", "» Clue: Country Words",
    "» Clue: Phone Country", "» Clue: TG Handles", "» Source Channel",
    "» Message ID", "» Message Date", "» Original Text", "» English Text",
]

# Columns a human must fill; highlighted so they are obvious in the sheet.
HUMAN_COLUMNS = {"Ecosystem (Geo)", "Vendor Type", "Geo of Acc."}

_HEADER_FONT = Font(bold=True)
_HUMAN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
_HELPER_FONT = Font(bold=True, italic=True, color="666666")


def _record_to_row(record):
    """Map one listing record to the template columns (geo columns stay blank)."""
    return {
        "Date of Finding": record.get("date_of_finding", ""),
        "URL": record.get("url", ""),
        "Vendor Type": "",  # human
        "Vendor Name": record.get("vendor_name", ""),
        "Ecosystem (Geo)": "",  # human
        "Account Type": record.get("account_type", ""),
        "Details": record.get("details", ""),
        "Age": record.get("age", ""),
        "Verified": record.get("verified", ""),
        "Farmed": record.get("farmed", ""),
        "Spend (if applicable)": record.get("spend", ""),
        "Payment Attached": record.get("payment_attached", ""),
        "PIN": record.get("pin", ""),
        "History": record.get("history", ""),
        "Geo of Acc.": "",  # human
        "Transfer Method": record.get("transfer_method", ""),
        "Cost": record.get("cost", ""),
        "No. Available": record.get("quantity", ""),
        "Payment Method (if detailed)": record.get("payment_method", ""),
        "Screenshot 1": "", "Screenshot 2": "", "Screenshot 3": "", "Screenshot 4": "",
    }


def _record_to_helpers(record):
    return {
        "» Confidence": record.get("confidence", ""),
        "» Needs Review": "YES" if record.get("needs_review") else "",
        "» Clue: Language": record.get("clue_language", ""),
        "» Clue: Country Words": ", ".join(record.get("clue_country_words", [])),
        "» Clue: Phone Country": record.get("clue_phone_country", ""),
        "» Clue: TG Handles": ", ".join(record.get("clue_handles", [])),
        "» Source Channel": record.get("source_channel", ""),
        "» Message ID": record.get("message_id", ""),
        "» Message Date": record.get("message_date", ""),
        "» Original Text": record.get("text_original", ""),
        "» English Text": record.get("text_english", ""),
    }


def build_workbook(records, invites=None):
    """Build the workbook: one 'Account Procurement' sheet + a 'Telegram Groups' sheet."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Account Procurement"

    all_columns = TEMPLATE_COLUMNS + HELPER_COLUMNS
    sheet.append(all_columns)
    for index, name in enumerate(all_columns, start=1):
        cell = sheet.cell(row=1, column=index)
        cell.font = _HELPER_FONT if name.startswith("»") else _HEADER_FONT
        if name in HUMAN_COLUMNS:
            cell.fill = _HUMAN_FILL

    for record in records:
        row = _record_to_row(record)
        helpers = _record_to_helpers(record)
        sheet.append([row[name] for name in TEMPLATE_COLUMNS] +
                     [helpers[name] for name in HELPER_COLUMNS])

    _build_groups_sheet(workbook, invites or [])
    return workbook


def _build_groups_sheet(workbook, invites):
    """Second tab for Telegram groups of interest / invite links."""
    groups = workbook.create_sheet("Telegram Groups")
    groups.append(["Invite Link", "Source Channel", "First Seen"])
    for cell in groups[1]:
        cell.font = _HEADER_FONT
    for invite in invites:
        groups.append([
            invite.get("invite_link", ""),
            invite.get("source_channel", ""),
            invite.get("first_seen", ""),
        ])


def save_sheet(records, out_path, invites=None):
    """Build and save the workbook. Returns the output Path."""
    out_path = Path(out_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    build_workbook(records, invites=invites).save(out_path)
    return out_path
